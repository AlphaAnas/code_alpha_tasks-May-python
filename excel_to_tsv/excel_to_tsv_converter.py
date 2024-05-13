#Task Automation with
# Python Scripts


#excel to tsv converter (tab separated values)

from openpyxl import load_workbook
import csv

file_name = "excelfile"    ## change the file name as per your excel file in input
book = load_workbook(filename=f"{file_name}.xlsx")

sheet = book.active

csv_lst = [] # list to collect the data from the excel file

for rows in sheet.iter_rows(values_only=True):
    csv_lst.append(list(rows))

#now write the data in a csv file
with open(f"{file_name}.tsv", 'w') as csv_obj:

    writer = csv.writer(csv_obj, delimiter=' ')
    for line in csv_lst:
        writer.writerow(line)

